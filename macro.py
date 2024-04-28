#   Copyright (c) 2024 Silvio Oliveira 
#   Contact: <silvio.oliveira347@gmail.com>; www.linkedin.com/in/silvio-oliveira-87155a203
 
import os
import datetime
import locale
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


# Cria uma tabela no excel
def criar_planilha(caminho_site, site):
    # Criar um novo workbook
    caminho_arquivo_xlsx = os.path.join(caminho_site, site + ".xlsx")
    wb = Workbook()
    ws = wb.active
 
    # Definir os cabeçalhos da tabela
    cabecalho = ["Item", "Descrição", "Quantidade"]
 
    # Preencher o cabeçalho da tabela
    for col, header in enumerate(cabecalho, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
 
    # Definir dados de exemplo
    dados = [
        ["--", "--", "--"],
        ["--", "--", "--"],
        ["--", "--", "--"]
    ]
 
    # Preencher os dados na tabela
    for linha, linha_dados in enumerate(dados, start=2):
        for col, dado in enumerate(linha_dados, start=1):
            ws.cell(row=linha, column=col, value=dado)
 
    # Definir estilos para as células
    for row in ws.iter_rows(min_row=1, max_row=len(dados) + 1, min_col=1, max_col=len(cabecalho)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                 top=Side(border_style='thin'), bottom=Side(border_style='thin'))
 
    # Ajustar a largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            if cell.value is not None:  # Verificar se a célula não está vazia
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2) * 3
        ws.column_dimensions[column].width = adjusted_width
 
    # Ajustar o zoom da planilha
    ws.sheet_view.zoomScale = 160
 
    # Salvar o arquivo
    wb.save(caminho_arquivo_xlsx)
   
    # Abrir o arquivo no Excel
    os.system(f'start excel.exe "{caminho_arquivo_xlsx}"')
 
# Cria o bloco de notas
def criar_txt(caminho_site, site):
    # Abre o arquivo .txt no modo de escrita
    caminho_arquivo_txt = os.path.join(caminho_site, site + ".txt")
    if not os.path.exists(caminho_arquivo_txt):
        with open(caminho_arquivo_txt, 'w') as arquivo:
            arquivo.write(f"SITE: {site} \n" +  
                          f"STATUS DO TSSR: \n" +
                          f"STATUS DO QRF: \n" +
                          f"STATUS DA LISTA DE MATERIAIS: \n\n" +
                          "---------------------------------------------------------------\n\n" +
                          "OBSERVAÇÕES:"
                          )
    os.system(f'start notepad.exe "{caminho_arquivo_txt}"')
 
# Função para determinar a região com base no código do estado
def determinar_regiao(site):
    estado = site[:2]  # Obtém os dois primeiros caracteres do código do site
    if estado == "BA":
        return "BA"
    elif estado in ["PR", "SC"]:
        return "SUL"
    elif estado in ["PB", "PE", "RN", "CE", "PI", "AL"]:
        return "NE"
    elif estado in ["DF", "RO", "MT", "GO", "AC", "TO", "MS"]:
        return "CO"
    elif estado == "MG":
        return "MG"
    elif estado == "ES":
        return "ES"
    elif estado in ["SI", "SM"]:
        return "SP"
    else:
        return "NOK"
 
# Define o idioma a ser utilizado
def selecionar_idioma(event=None):
    global msg
    idioma_selecionado = combobox_idioma.get()
    msg = CriaMensagens(idioma_selecionado)
    CriaTextosWidgets(msg)
    CriarCaixas(msg)
   
def CriaMensagens(idioma_selecionado):
    if idioma_selecionado == "Português":
        mensagens = [
            "Erro",
            "Por favor, preencha todos os campos.",
            "Por favor, preencha os campos corretamente.",
            "Ocorreu um erro.",
            "Script de Criação de Pasta do Site",
            "Defina o diretório para armazenar os sites:",
            "Coloque o nome do seu site:",
            "Gerar planilha",
            "Gerar bloco de notas",
            "Idiomas",
            "Crie sua pasta!",
            "Item",
            "Descrição",
            "Quantidade",
            "Gerar planilha",
            "Gerar bloco de notas",
            "Abrir pasta após criar",
            "A pasta já está criada, nada feito."
        ]
    elif idioma_selecionado == "Inglês":
        mensagens = [
            "Error",
            "Please fill in all fields.",
            "Please fill in the fields correctly.",
            "An error occurred.",
            "Site Folder Creation Script",
            "Define the directory to store the sites:",
            "Enter your site's name:",
            "Generate spreadsheet",
            "Generate notepad",
            "Languages",
            "Create your folder!",
            "Item",
            "Description",
            "Quantity",
            "Generate spreadsheet",
            "Generate notepad",
            "Open folder after creating",
            "Folder already exists, nothing done."
        ]
    elif idioma_selecionado == "Espanhol":
        mensagens = [
            "Error",
            "Por favor, complete todos los campos.",
            "Por favor, complete los campos correctamente.",
            "Se ha producido un error.",
            "Script de Creación de Carpetas de Sitios",
            "Defina el directorio para almacenar los sitios:",
            "Ingrese el nombre de su sitio:",
            "Generar hoja de cálculo",
            "Generar bloc de notas",
            "Idiomas",
            "¡Cree su carpeta!",
            "Artículo",
            "Descripción",
            "Cantidad",
            "Generar hoja de cálculo",
            "Generar bloc de notas",
            "Abrir carpeta después de crear",
            "La carpeta ya existe, no se realizó ninguna acción."
        ]
    return mensagens
 
# Retorna o nome do usuário do PC
def get_username():
    return os.getenv('USERNAME')
   
# Função raíz
def criarpasta(event=None):
    try:
        site = entrada_site.get().upper().strip()   # Site do campo de entrada (transforma tudo em maiúsculo)
        diretorio = entrada_diretorio.get().lower().strip()   # Diretório do campo de entrada (transforma tudo em minúsculo)
               
        if not diretorio or not site:
            messagebox.showerror(msg[0], msg[1])
            return
 
        if len(diretorio) < 8:
            messagebox.showerror(msg[0], msg[2])
            return
 
        # Verificando se o diretório é "Downloads" ou "Desktop"
        if diretorio.lower() == "downloads":
            diretorio = os.path.join("C:\\Users", get_username(), "Downloads")
        elif diretorio.lower() == "desktop":
            diretorio = os.path.join("C:\\Users", get_username(), "Desktop")
 
        # Configurando o locale para português do Brasil
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
   
        # Obtendo o nome do mês atual em português
        mes = datetime.datetime.now().strftime('%B').capitalize()
 
        # Determinar regiao
        regiao = determinar_regiao(site)
   
        # Pasta da regional
        caminho_pasta_regional = os.path.join(diretorio, "Sites", mes, regiao)
        if not os.path.exists(caminho_pasta_regional) or not os.path.isdir(caminho_pasta_regional):
            os.makedirs(caminho_pasta_regional)
        else:
            messagebox.showerror("Erro", msg[17])
            os.startfile(caminho_pasta_regional)
            return
 
        # Direcionando a pasta do site
        caminho_site = os.path.join(caminho_pasta_regional, site)
        if not os.path.exists(caminho_site) or not os.path.isdir(caminho_site):
            os.mkdir(caminho_site)
        else:
            messagebox.showerror("Erro", msg[17])
            os.startfile(caminho_site)
            return

        # Criando e mostrando arquivos
        if opcao1_var.get():
            criar_planilha(caminho_site, site)  # Criar planilha se a opção estiver selecionada
        if opcao2_var.get():
            criar_txt(caminho_site, site)  # Criar bloco de notas se a opção estiver selecionada
        
        # Abrir a pasta se a opção estiver selecionada
        if opcao3_var.get():
            os.startfile(caminho_site)
 
    except Exception as e:
        messagebox.showerror(msg[3])
 
 
############################################front#############################################

# Alternar mensagens que ficam acima dos campos do texto
def CriaTextosWidgets(msg):
    global texto
    global texto2
    global botao
 
    try:
        texto.destroy()
        texto2.destroy()
        botao.destroy()
    except:
        pass

    texto = tk.Label(janela, text=msg[5], pady=20, font=fonte_texto)
    texto.grid(column=0, row=0, sticky="n", padx=10)
    texto2 = tk.Label(janela, text=msg[6], pady=20, font=fonte_texto)
    texto2.grid(column=0, row=2, sticky="n", padx=10)
   
    # Botão para criar a pasta
    botao = tk.Button(janela, text=msg[10], command=criarpasta, font=fonte_texto)
    botao.grid(column=0, row=4, sticky="n", padx=10, pady=20)
      
# Cria os campos de texto para diretório e site    
def CamposDeTexto():
    global entrada_diretorio
    global entrada_site

    # Campos de entrada
    entrada_diretorio = tk.Entry(janela, width=50, font=fonte_texto)
    entrada_diretorio.grid(column=0, row=1, sticky="n", padx=10)
    entrada_diretorio.bind("<Return>", criarpasta)  # Vincular o evento/tecla "Enter" à função criarpasta
 
    entrada_site = tk.Entry(janela, width=50, font=fonte_texto)
    entrada_site.grid(column=0, row=3, sticky="n", padx=10)
    entrada_site.bind("<Return>", criarpasta)  # Vincular o evento/tecla "Enter" à função criarpasta

opcao1 = None
opcao2 = None
opcao3 = None
# Cria caixas para selecionar algumas opções
def CriarCaixas(msg):
    global opcao1
    global opcao2
    global opcao3
    global opcao1_var
    global opcao2_var
    global opcao3_var

    # Remover widgets antigos, se existirem
    if opcao1:
        opcao1.grid_forget()
        opcao1.destroy()
    if opcao2:
        opcao2.grid_forget()
        opcao2.destroy()
    if opcao3:
        opcao3.grid_forget()
        opcao3.destroy()

    # Variáveis de controle para as opções
    opcao1_var = tk.BooleanVar(value=True)  # Inicialmente marcado
    opcao2_var = tk.BooleanVar(value=True)  # Inicialmente marcado
    opcao3_var = tk.BooleanVar(value=True)  # Inicialmente marcado
    
    # Checkbutton para selecionar a opção "Gerar planilha"
    opcao1 = ttk.Checkbutton(janela, text=msg[14], variable=opcao1_var)
    opcao1.grid(row=1, column=1, sticky="w", padx=(35, 10), pady=(5, 0))
    
    # Checkbutton para selecionar a opção "Gerar bloco de notas"
    opcao2 = ttk.Checkbutton(janela, text=msg[15], variable=opcao2_var)
    opcao2.grid(row=2, column=1, sticky="w", padx=(35, 10), pady=(5, 0))
    
    # Checkbutton para selecionar a opção "Abrir pasta após criar"
    opcao3 = ttk.Checkbutton(janela, text=msg[16], variable=opcao3_var)
    opcao3.grid(row=3, column=1, sticky="w", padx=(35, 10))

# Função para centralizar a janela na tela do usuário
def centralizar_janela(janela):
    janela.update_idletasks()
    largura = janela.winfo_width()
    altura = janela.winfo_height()
    x = (janela.winfo_screenwidth() // 2) - (largura // 2)
    y = (janela.winfo_screenheight() // 2) - (altura // 2)
    janela.geometry('{}x{}+{}+{}'.format(largura, altura, x, y))

# Criando a janela
janela = tk.Tk()
janela.title("Script For Organization")

# Definindo o estilo do tema
style = ttk.Style(janela)
janela.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

# Definindo a fonte para todos os textos
fonte_texto = ("Helvetica", 10, "bold")

# Escrevendo "Português" na barra de idiomas
idioma_selecionado = tk.StringVar(value="Português")

# Opções de idioma
idiomas = ["Português", "Inglês", "Espanhol"]

# Combobox para selecionar o idioma
combobox_idioma = ttk.Combobox(janela, textvariable=idioma_selecionado, values=idiomas, font=fonte_texto, width=15)
combobox_idioma.grid(row=0, column=1, padx=30, pady=10)
combobox_idioma.bind("<<ComboboxSelected>>", selecionar_idioma)

# Inicializa em PT
selecionar_idioma(idiomas[0])

# Campos de Texto
CamposDeTexto()

# Criar Caixas
CriarCaixas(msg)

# Aumentando a janela
janela.geometry("650x260")

# Centralizando a janela
centralizar_janela(janela)

# Mantendo a janela aberta
janela.mainloop()